import pandas as pd 
import kumo_with_future as kwf
from openpyxl import load_workbook


def fillProfit(pos, profit,entries):
    bufferFile_pd = pd.read_excel('NIFTY200.xlsx')
    bufferFile_ = load_workbook('NIFTY200.xlsx')
    bufferFile_sheet_ = bufferFile_.active
    bufferFile_sheet_.cell(row=pos, column=2).value = profit
    bufferFile_sheet_.cell(row=pos, column=3).value = entries
    bufferFile_.save('NIFTY200.xlsx')



if __name__ == "__main__":
    
    ma = pd.read_excel('NIFTY200.xlsx')
    i = 2
    for symbol in ma['Symbol']:
        data = kwf.fetchTickerData2(symbol)
        kwf.initialise()
        data = kwf.trading(data)
        fillProfit(i,sum(kwf.profit),len(kwf.profit))
        i+=1
        